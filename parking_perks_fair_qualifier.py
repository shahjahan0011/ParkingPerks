# =============================================================================
# PARKING PERKS — Fair Monthly Qualifier Report
# =============================================================================
#
# This version is designed to avoid silently missing people.
#
# Key behavior:
#   1. Official qualification still uses the policy threshold:
#        N+ UNIQUE qualifying days, valid payment/permit, no citations.
#   2. It calculates qualification using TWO evidence methods:
#        A) daily-span method:
#           first read of the day -> last read of the day
#        B) session method:
#           continuous sessions split by a configurable gap threshold
#   3. Default official method is daily-span because the stated rule is
#      "10+ separate days", not "10+ continuous sessions".
#   4. The report also includes Review Candidates so borderline or partial-data
#      cases are not lost.
#   5. If the reads file has fewer unique dates than --min-visits, the script
#      warns clearly because official qualification is impossible unless you
#      explicitly pass --allow-partial-month.
#
# Example:
#   python parking_perks_fair_monthly.py \
#     --reads "April Plate Reads.xlsx" \
#     --payments "April Payments.csv" \
#     --citations "Citations April.xls" \
#     --output "Parking_Perks_April_2026.xlsx"
#
# Testing partial data:
#   python parking_perks_fair_monthly.py \
#     --reads "April Plate Reads.xlsx" \
#     --payments "April Payments.csv" \
#     --citations "Citations April.xls" \
#     --min-visits 5 \
#     --output "Parking_Perks_Test.xlsx"
#
# If leadership explicitly wants proportional/partial-month official results:
#   add --allow-partial-month
# =============================================================================

import argparse
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# NORMALIZATION
# =============================================================================

def normalise_reads_plate(raw) -> str:
    if pd.isna(raw):
        return ""
    return str(raw).strip().upper().replace(" ", "")


def normalise_payments_plate(raw) -> str:
    if pd.isna(raw):
        return ""

    s = str(raw).strip().upper()

    # Handles values like ="SK041H"
    if s.startswith('="') and s.endswith('"'):
        s = s[2:-1]
    else:
        s = s.lstrip('="').rstrip('"')

    return s.replace(" ", "")


def normalise_citations_plate(raw) -> str:
    if pd.isna(raw):
        return ""

    s = str(raw).strip().upper()

    if not s or s == "LICENSE #":
        return ""

    parts = [p.strip() for p in s.split("-")]

    # Standard citation format: BC-SK041H-NA
    if len(parts) == 3 and parts[1]:
        return parts[1].replace(" ", "")

    # Already plain plate
    if len(parts) == 1:
        return parts[0].replace(" ", "")

    # Defensive fallback for odd values like BC-SK-041H-NA
    if len(parts) > 3 and parts[0] and parts[-1]:
        return "".join(parts[1:-1]).replace(" ", "")

    return ""


# =============================================================================
# HELPERS
# =============================================================================

def find_column(df: pd.DataFrame, keywords: list[str]) -> str | None:
    return next(
        (col for col in df.columns if any(k in str(col).lower() for k in keywords)),
        None
    )


def excel_engine_for(path: str) -> str | None:
    suffix = Path(path).suffix.lower()
    if suffix == ".xls":
        return "xlrd"
    if suffix == ".xlsx":
        return "openpyxl"
    return None


def require_columns(df: pd.DataFrame, required_cols: list[str], file_label: str) -> None:
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        raise ValueError(
            f"{file_label} is missing required column(s): {missing}. "
            f"Available columns: {', '.join(map(str, df.columns))}"
        )


# =============================================================================
# LOADERS
# =============================================================================

def load_plate_reads(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, header=1)
    df.columns = df.columns.str.strip()

    df = df.rename(columns={
        "Local time (PDT)": "local_time",
        "Plate number": "plate_raw",
    })

    require_columns(df, ["local_time", "plate_raw"], "Plate reads file")

    df["plate"] = df["plate_raw"].apply(normalise_reads_plate)
    df = df[df["plate"] != ""]

    df["local_time"] = pd.to_datetime(
        df["local_time"],
        format="%m/%d/%Y, %I:%M:%S %p",
        errors="coerce",
    )

    df = df.dropna(subset=["local_time"])
    df = df.drop_duplicates(subset=["plate", "local_time"], keep="first")

    return df[["plate", "local_time"]].copy()


def load_payments(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str)
    df.columns = df.columns.str.strip()

    plate_col = find_column(df, ["license plate", "licence plate", "plate"])
    if not plate_col:
        raise ValueError(
            "Payments file must contain a plate column. "
            f"Available columns: {', '.join(map(str, df.columns))}"
        )

    df["plate"] = df[plate_col].apply(normalise_payments_plate)
    df = df[df["plate"] != ""]

    return df[["plate"]].drop_duplicates().copy()


def load_citations(path: str) -> pd.DataFrame:
    engine = excel_engine_for(path)

    try:
        df = pd.read_excel(path, engine=engine, header=9)
    except ImportError as exc:
        raise ImportError(
            "Reading .xls citation files requires xlrd. Install it with: pip install xlrd"
        ) from exc

    df.columns = df.columns.str.strip()

    plate_col = find_column(df, ["license #", "licence #", "license", "licence", "plate"])
    if not plate_col:
        raise ValueError(
            "Citations file must contain a citation plate column. "
            f"Available columns: {', '.join(map(str, df.columns))}"
        )

    df["plate"] = df[plate_col].apply(normalise_citations_plate)
    df = df[df["plate"] != ""]

    return df[["plate"]].drop_duplicates().copy()


def load_permits(path: str) -> pd.DataFrame:
    suffix = Path(path).suffix.lower()

    if suffix == ".csv":
        df = pd.read_csv(path, dtype=str)
    else:
        df = pd.read_excel(path, dtype=str, engine=excel_engine_for(path))

    df.columns = df.columns.str.strip()

    plate_col = find_column(df, ["plate", "licence", "license"])
    permit_col = find_column(df, ["permit"])
    name_col = find_column(df, ["name"])

    if not plate_col:
        return pd.DataFrame(columns=["plate"])

    out = pd.DataFrame()
    out["plate"] = df[plate_col].apply(normalise_payments_plate)

    if permit_col:
        out["permit_number"] = df[permit_col].fillna("").astype(str)

    if name_col:
        out["name"] = df[name_col].fillna("").astype(str)

    out = out[out["plate"] != ""]
    return out.drop_duplicates(subset=["plate"]).copy()


# =============================================================================
# VISIT METHODS
# =============================================================================

def compute_daily_span_summary(reads: pd.DataFrame, min_hours: float) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Daily-span method:
    For each plate/date, compare first read and last read on that same date.
    This preserves the "separate days" policy and avoids missing people because
    cameras did not continuously observe them.
    """
    daily = (
        reads.assign(visit_date=reads["local_time"].dt.date)
        .groupby(["plate", "visit_date"])
        .agg(
            first_read=("local_time", "min"),
            last_read=("local_time", "max"),
            daily_read_count=("local_time", "count"),
        )
        .reset_index()
    )

    daily["daily_span_hrs"] = (
        daily["last_read"] - daily["first_read"]
    ).dt.total_seconds() / 3600

    qualifying_daily = daily[daily["daily_span_hrs"] >= min_hours].copy()

    if qualifying_daily.empty:
        return pd.DataFrame(columns=[
            "plate",
            "daily_qualifying_days",
            "avg_daily_span_hrs",
            "first_daily_qualifying_date",
            "last_daily_qualifying_date",
        ]), qualifying_daily

    summary = (
        qualifying_daily.groupby("plate")
        .agg(
            daily_qualifying_days=("visit_date", "nunique"),
            avg_daily_span_hrs=("daily_span_hrs", "mean"),
            first_daily_qualifying_date=("visit_date", "min"),
            last_daily_qualifying_date=("visit_date", "max"),
        )
        .reset_index()
    )

    summary["avg_daily_span_hrs"] = summary["avg_daily_span_hrs"].round(1)

    return summary, qualifying_daily


def compute_session_summary(
    reads: pd.DataFrame,
    min_hours: float,
    session_gap_minutes: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Session method:
    Split reads into sessions when the gap between two reads exceeds the
    threshold. Useful as supporting evidence, but can undercount if camera
    reads are sparse.
    """
    reads = reads.sort_values(["plate", "local_time"]).copy()

    reads["prev_time"] = reads.groupby("plate")["local_time"].shift(1)
    reads["gap_mins"] = (
        reads["local_time"] - reads["prev_time"]
    ).dt.total_seconds() / 60

    reads["is_new_session"] = (
        reads["gap_mins"].isna() |
        (reads["gap_mins"] > session_gap_minutes)
    )

    reads["session_id"] = reads.groupby("plate")["is_new_session"].cumsum()

    sessions = (
        reads.groupby(["plate", "session_id"])
        .agg(
            session_start=("local_time", "min"),
            session_end=("local_time", "max"),
            session_read_count=("local_time", "count"),
        )
        .reset_index()
    )

    sessions["session_duration_hrs"] = (
        sessions["session_end"] - sessions["session_start"]
    ).dt.total_seconds() / 3600

    sessions["session_start_date"] = sessions["session_start"].dt.date

    qualifying_sessions = sessions[sessions["session_duration_hrs"] >= min_hours].copy()

    if qualifying_sessions.empty:
        return pd.DataFrame(columns=[
            "plate",
            "session_qualifying_days",
            "qualifying_sessions",
            "avg_session_hrs",
        ]), qualifying_sessions

    summary = (
        qualifying_sessions.groupby("plate")
        .agg(
            session_qualifying_days=("session_start_date", "nunique"),
            qualifying_sessions=("session_id", "count"),
            avg_session_hrs=("session_duration_hrs", "mean"),
        )
        .reset_index()
    )

    summary["avg_session_hrs"] = summary["avg_session_hrs"].round(1)

    return summary, qualifying_sessions


def build_method_comparison(
    daily_summary: pd.DataFrame,
    session_summary: pd.DataFrame,
) -> pd.DataFrame:
    comparison = pd.merge(
        daily_summary,
        session_summary,
        on="plate",
        how="outer",
    ).fillna({
        "daily_qualifying_days": 0,
        "avg_daily_span_hrs": 0,
        "session_qualifying_days": 0,
        "qualifying_sessions": 0,
        "avg_session_hrs": 0,
    })

    for col in ["daily_qualifying_days", "session_qualifying_days", "qualifying_sessions"]:
        comparison[col] = comparison[col].astype(int)

    comparison["method_difference_days"] = (
        comparison["daily_qualifying_days"] - comparison["session_qualifying_days"]
    )

    return comparison


# =============================================================================
# REPORT WRITER
# =============================================================================

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F4E79")
BODY_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", start_color="EBF3FB")
GREEN_FILL = PatternFill("solid", start_color="E2EFDA")
GREEN_FONT = Font(name="Arial", size=10, color="375623")
WARNING_FILL = PatternFill("solid", start_color="FFF2CC")
WARNING_FONT = Font(name="Arial", size=10, color="7F6000", bold=True)
THIN_SIDE = Side(style="thin", color="B8CCE4")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def style_header_row(ws, row_num: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_table_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    subtitle: str,
    df: pd.DataFrame,
    column_order: list[str],
    column_labels: list[str],
    widths: list[int],
) -> None:
    ws = wb.create_sheet(sheet_name)

    total_cols = max(1, len(column_order))
    last_col = get_column_letter(total_cols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    for col_num, label in enumerate(column_labels, 1):
        ws.cell(row=3, column=col_num, value=label)

    style_header_row(ws, 3, len(column_order))
    ws.freeze_panes = "A4"

    if df.empty:
        ws.cell(row=4, column=1, value="No rows found.")
        ws.cell(row=4, column=1).font = BODY_FONT
    else:
        df_to_write = df.copy()
        for col in column_order:
            if col not in df_to_write.columns:
                df_to_write[col] = ""

        df_to_write = df_to_write[column_order]

        for row_idx, (_, row) in enumerate(df_to_write.iterrows(), start=4):
            fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()

            for col_num, col_name in enumerate(column_order, 1):
                value = row[col_name]
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                cell.font = BODY_FONT
                cell.fill = fill
                cell.border = CELL_BORDER
                cell.alignment = Alignment(vertical="center")

    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width


def write_report(
    output_path: str,
    month_label: str,
    official: pd.DataFrame,
    review: pd.DataFrame,
    comparison: pd.DataFrame,
    summary_stats: dict,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    official_cols = [
        "plate",
        "name",
        "permit_number",
        "daily_qualifying_days",
        "session_qualifying_days",
        "qualifying_sessions",
        "avg_daily_span_hrs",
        "avg_session_hrs",
        "payment_source",
    ]

    official_labels = [
        "Plate",
        "Name",
        "Permit Number",
        "Daily Qualifying Days",
        "Session Qualifying Days",
        "Qualifying Sessions",
        "Avg Daily Span (hrs)",
        "Avg Session (hrs)",
        "Payment Source",
    ]

    write_table_sheet(
        wb,
        "Official Qualifiers",
        f"Parking Perks — Official Qualifiers | {month_label}",
        (
            f"Criteria: {summary_stats['effective_min_visits']}+ unique qualifying days, "
            f"{summary_stats['min_hours']}+ observed hr, valid payment/permit, no citations. "
            f"Official method: {summary_stats['official_method']}."
        ),
        official,
        official_cols,
        official_labels,
        [16, 26, 18, 22, 22, 20, 20, 18, 16],
    )

    review_cols = [
        "plate",
        "review_reason",
        "daily_qualifying_days",
        "session_qualifying_days",
        "qualifying_sessions",
        "avg_daily_span_hrs",
        "avg_session_hrs",
        "has_payment",
        "has_permit",
        "has_citation",
    ]

    review_labels = [
        "Plate",
        "Review Reason",
        "Daily Qualifying Days",
        "Session Qualifying Days",
        "Qualifying Sessions",
        "Avg Daily Span (hrs)",
        "Avg Session (hrs)",
        "Has Payment",
        "Has Permit",
        "Has Citation",
    ]

    write_table_sheet(
        wb,
        "Review Candidates",
        f"Parking Perks — Review Candidates | {month_label}",
        "These plates should be reviewed so nobody is missed because of partial data or method disagreement.",
        review,
        review_cols,
        review_labels,
        [16, 42, 22, 22, 20, 20, 18, 14, 14, 14],
    )

    comparison_cols = [
        "plate",
        "daily_qualifying_days",
        "session_qualifying_days",
        "qualifying_sessions",
        "method_difference_days",
        "avg_daily_span_hrs",
        "avg_session_hrs",
        "has_payment",
        "has_permit",
        "has_citation",
    ]

    comparison_labels = [
        "Plate",
        "Daily Qualifying Days",
        "Session Qualifying Days",
        "Qualifying Sessions",
        "Daily - Session Difference",
        "Avg Daily Span (hrs)",
        "Avg Session (hrs)",
        "Has Payment",
        "Has Permit",
        "Has Citation",
    ]

    write_table_sheet(
        wb,
        "Method Comparison",
        f"Parking Perks — Method Comparison | {month_label}",
        "Shows where daily-span and session methods disagree.",
        comparison,
        comparison_cols,
        comparison_labels,
        [16, 22, 22, 20, 24, 20, 18, 14, 14, 14],
    )

    # Summary sheet
    ws = wb.create_sheet("Processing Summary")
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 28

    ws["A1"] = "Parking Perks — Processing Summary"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    rows = [
        ("Month label", month_label),
        ("Run date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        (None, None),
        ("-- Coverage --", None),
        ("Reads date range", summary_stats["date_range"]),
        ("Unique dates in reads file", summary_stats["coverage_days"]),
        ("Policy minimum unique qualifying days", summary_stats["min_visits"]),
        ("Effective official minimum used", summary_stats["effective_min_visits"]),
        ("Partial month override used", summary_stats["allow_partial_month"]),
        (None, None),
        ("-- Inputs --", None),
        ("Cleaned plate read rows", summary_stats["read_rows"]),
        ("Unique plates in reads", summary_stats["total_plates"]),
        ("Unique paid plates", summary_stats["payment_plates"]),
        ("Unique cited plates", summary_stats["citation_plates"]),
        ("Unique permit plates", summary_stats["permit_plates"]),
        (None, None),
        ("-- Visit evidence --", None),
        ("Minimum hours per qualifying visit", summary_stats["min_hours"]),
        ("Session gap threshold", f"{summary_stats['session_gap_minutes']} minutes"),
        ("Daily-span qualifying day rows", summary_stats["daily_qualifying_rows"]),
        ("Session qualifying rows", summary_stats["session_qualifying_rows"]),
        (None, None),
        ("-- Results --", None),
        ("Official method", summary_stats["official_method"]),
        ("Stage 1 official candidates before payment/citations", summary_stats["stage1"]),
        ("Stage 2 with payment/permit", summary_stats["stage2"]),
        ("Removed because of citation", summary_stats["removed_citations"]),
        ("FINAL OFFICIAL QUALIFIERS", summary_stats["final"]),
        ("Review candidate rows", summary_stats["review_count"]),
    ]

    for row_idx, (label, value) in enumerate(rows, start=2):
        is_section = bool(label and label.startswith("--"))
        is_final = label == "FINAL OFFICIAL QUALIFIERS"

        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(
            name="Arial",
            size=10,
            bold=(is_section or is_final),
            color="1F4E79" if is_final else "000000",
        )

        if value is not None:
            value_cell = ws.cell(row=row_idx, column=2, value=value)
            value_cell.font = Font(
                name="Arial",
                size=10,
                bold=is_final,
                color="1F4E79" if is_final else "000000",
            )
            value_cell.alignment = Alignment(horizontal="right")

    if summary_stats["coverage_days"] < summary_stats["min_visits"]:
        warning_row = len(rows) + 4
        ws.merge_cells(start_row=warning_row, start_column=1, end_row=warning_row, end_column=2)
        warning_cell = ws.cell(row=warning_row, column=1)
        warning_cell.value = (
            "WARNING: The reads file has fewer unique dates than the policy minimum. "
            "Official qualification is impossible unless --allow-partial-month is used."
        )
        warning_cell.fill = WARNING_FILL
        warning_cell.font = WARNING_FONT
        warning_cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[warning_row].height = 42

    wb.save(output_path)


# =============================================================================
# MAIN
# =============================================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Parking Perks — fair monthly qualifier report"
    )

    parser.add_argument("--reads", required=True, help="Plate reads file (.xlsx)")
    parser.add_argument("--payments", required=True, help="Payments file (.csv)")
    parser.add_argument("--citations", required=True, help="Citations file (.xls or .xlsx)")
    parser.add_argument("--permits", default=None, help="Optional permit holders file (.csv or .xlsx)")
    parser.add_argument("--output", default="Parking_Perks_Report.xlsx", help="Output Excel filename")
    parser.add_argument("--min-visits", type=int, default=10, help="Policy minimum unique qualifying days")
    parser.add_argument("--min-hours", type=float, default=1.0, help="Minimum observed hours")
    parser.add_argument("--session-gap-minutes", type=int, default=120, help="Gap after which a new session starts")
    parser.add_argument(
        "--official-method",
        choices=["daily-span", "session", "either"],
        default="daily-span",
        help=(
            "daily-span: official based on first-to-last read per day; "
            "session: official based on session method; "
            "either: official if either method meets the threshold."
        ),
    )
    parser.add_argument(
        "--allow-partial-month",
        action="store_true",
        help=(
            "If reads file has fewer unique dates than --min-visits, use the number "
            "of available unique dates as the official minimum. Use only if approved."
        ),
    )
    parser.add_argument(
        "--review-buffer-days",
        type=int,
        default=2,
        help="Include paid/uncited review candidates within this many days of the effective threshold.",
    )

    args = parser.parse_args()

    month_label = (
        os.path.basename(args.reads)
        .replace("_", " ")
        .replace(".xlsx", "")
        .replace(".xls", "")
    )

    print("\nLoading and cleaning files...")
    reads = load_plate_reads(args.reads)
    payments = load_payments(args.payments)
    citations = load_citations(args.citations)
    permits = load_permits(args.permits) if args.permits else pd.DataFrame(columns=["plate"])

    coverage_days = reads["local_time"].dt.date.nunique()
    date_range = f"{reads['local_time'].min()} to {reads['local_time'].max()}"
    total_plates = reads["plate"].nunique()

    print(f"Plate reads: {len(reads):,} cleaned rows | {total_plates:,} unique plates")
    print(f"Date range:  {date_range}")
    print(f"Coverage:    {coverage_days} unique dates")
    print(f"Payments:    {len(payments):,} unique plates")
    print(f"Citations:   {len(citations):,} unique plates")
    print(f"Permits:     {len(permits):,} unique plates")

    if coverage_days < args.min_visits:
        print(
            f"\nWARNING: The reads file only has {coverage_days} unique dates, "
            f"but the policy threshold is {args.min_visits}. "
            f"Official qualification is impossible without --allow-partial-month."
        )

    effective_min_visits = args.min_visits
    if args.allow_partial_month and coverage_days < args.min_visits:
        effective_min_visits = coverage_days
        print(
            f"Partial-month override enabled. Effective official minimum changed to "
            f"{effective_min_visits} unique qualifying days."
        )

    print("\nComputing daily-span evidence...")
    daily_summary, qualifying_daily = compute_daily_span_summary(reads, args.min_hours)
    print(f"Daily-span qualifying day rows: {len(qualifying_daily):,}")

    print("\nComputing session evidence...")
    session_summary, qualifying_sessions = compute_session_summary(
        reads,
        args.min_hours,
        args.session_gap_minutes,
    )
    print(f"Session qualifying rows: {len(qualifying_sessions):,}")

    comparison = build_method_comparison(daily_summary, session_summary)

    paid_plates = set(payments["plate"])
    permit_plates = set(permits["plate"]) if not permits.empty else set()
    cited_plates = set(citations["plate"])

    comparison["has_payment"] = comparison["plate"].isin(paid_plates)
    comparison["has_permit"] = comparison["plate"].isin(permit_plates)
    comparison["has_citation"] = comparison["plate"].isin(cited_plates)
    comparison["valid_payment"] = comparison["has_payment"] | comparison["has_permit"]

    if args.official_method == "daily-span":
        comparison["official_visit_days"] = comparison["daily_qualifying_days"]
    elif args.official_method == "session":
        comparison["official_visit_days"] = comparison["session_qualifying_days"]
    else:
        comparison["official_visit_days"] = comparison[
            ["daily_qualifying_days", "session_qualifying_days"]
        ].max(axis=1)

    stage1 = comparison[comparison["official_visit_days"] >= effective_min_visits].copy()
    stage2 = stage1[stage1["valid_payment"]].copy()
    official = stage2[~stage2["has_citation"]].copy()

    removed_citations = len(stage2) - len(official)

    official["payment_source"] = official.apply(
        lambda row: "Permit" if row["has_permit"] else "Payment",
        axis=1,
    )

    if not permits.empty:
        permit_cols = ["plate"]
        if "name" in permits.columns:
            permit_cols.append("name")
        if "permit_number" in permits.columns:
            permit_cols.append("permit_number")
        official = official.merge(permits[permit_cols], on="plate", how="left")

    if "name" not in official.columns:
        official["name"] = ""
    if "permit_number" not in official.columns:
        official["permit_number"] = ""

    official[["name", "permit_number"]] = official[["name", "permit_number"]].fillna("")

    official = official.sort_values(
        ["official_visit_days", "daily_qualifying_days", "session_qualifying_days"],
        ascending=[False, False, False],
    ).reset_index(drop=True)

    # Review candidates:
    # paid/uncited plates that are close to the threshold, disagree by method,
    # or are candidates under partial coverage.
    review_floor = max(1, effective_min_visits - args.review_buffer_days)

    review = comparison[
        comparison["valid_payment"] &
        ~comparison["has_citation"] &
        (
            (comparison["daily_qualifying_days"] >= review_floor) |
            (comparison["session_qualifying_days"] >= review_floor) |
            (comparison["method_difference_days"].abs() >= args.review_buffer_days)
        )
    ].copy()

    official_plates = set(official["plate"])
    review = review[~review["plate"].isin(official_plates)].copy()

    def review_reason(row):
        reasons = []

        if coverage_days < args.min_visits:
            reasons.append("partial reads file")

        if row["daily_qualifying_days"] >= review_floor:
            reasons.append("near threshold by daily-span")

        if row["session_qualifying_days"] >= review_floor:
            reasons.append("near threshold by session")

        if abs(row["method_difference_days"]) >= args.review_buffer_days:
            reasons.append("daily/session method disagreement")

        return "; ".join(reasons)

    if not review.empty:
        review["review_reason"] = review.apply(review_reason, axis=1)
        review = review.sort_values(
            ["daily_qualifying_days", "session_qualifying_days", "qualifying_sessions"],
            ascending=[False, False, False],
        ).reset_index(drop=True)
    else:
        review["review_reason"] = ""

    comparison = comparison.sort_values(
        ["daily_qualifying_days", "session_qualifying_days", "qualifying_sessions"],
        ascending=[False, False, False],
    ).reset_index(drop=True)

    print("\nResults:")
    print(f"Stage 1 official candidates before payment/citations: {len(stage1):,}")
    print(f"Stage 2 with payment/permit: {len(stage2):,}")
    print(f"Removed because of citation: {removed_citations:,}")
    print(f"Final official qualifiers: {len(official):,}")
    print(f"Review candidates: {len(review):,}")

    summary_stats = {
        "date_range": date_range,
        "coverage_days": coverage_days,
        "min_visits": args.min_visits,
        "effective_min_visits": effective_min_visits,
        "allow_partial_month": str(args.allow_partial_month),
        "read_rows": len(reads),
        "total_plates": total_plates,
        "payment_plates": len(payments),
        "citation_plates": len(citations),
        "permit_plates": len(permits),
        "min_hours": args.min_hours,
        "session_gap_minutes": args.session_gap_minutes,
        "daily_qualifying_rows": len(qualifying_daily),
        "session_qualifying_rows": len(qualifying_sessions),
        "official_method": args.official_method,
        "stage1": len(stage1),
        "stage2": len(stage2),
        "removed_citations": removed_citations,
        "final": len(official),
        "review_count": len(review),
    }

    print("\nWriting Excel report...")
    write_report(args.output, month_label, official, review, comparison, summary_stats)
    print(f"Report saved: {args.output}")
    print("Done.\n")


if __name__ == "__main__":
    main()
