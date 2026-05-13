# =============================================================================
# PARKING PERKS — Monthly Qualifier Report
# =============================================================================
#
# WHAT THIS SCRIPT DOES:
#   1. Reads three files: plate camera reads, payments, citations
#   2. Finds plates that visited campus on 10+ separate days,
#      each stay lasting more than 1 hour (first read to last read that day)
#   3. Keeps only plates that have a valid payment or permit on file
#   4. Removes any plate that received a citation that month
#   5. Writes a formatted Excel report with the qualifying plates
#
# HOW TO RUN IT (full month):
#   python parking_perks.py \
#     --reads     "April_Plate_Reads.xlsx" \
#     --payments  "April_Payments.csv" \
#     --citations "Citations_April.xls" \
#     --output    "Parking_Perks_April_2026.xlsx"
#
# HOW TO RUN IT (testing with partial data):
#   python parking_perks.py \
#     --reads     "April_Plate_Reads.xlsx" \
#     --payments  "April_Payments.csv" \
#     --citations "Citations_April.xls" \
#     --min-visits 4 \
#     --output    "Parking_Perks_TEST.xlsx"
#
# ADDING PERMITS (when available):
#   python parking_perks.py \
#     --reads     "April_Plate_Reads.xlsx" \
#     --payments  "April_Payments.csv" \
#     --citations "Citations_April.xls" \
#     --permits   "Active_Permits.csv" \
#     --output    "Parking_Perks_April_2026.xlsx"
#
# REQUIREMENTS:
#   pip install pandas openpyxl xlrd
# =============================================================================

import argparse
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# SECTION 1 — PLATE NORMALISATION
# =============================================================================
#
# Each file stores the same physical plate number in a different format.
# We write one small function per file so each only does exactly what
# that file needs — nothing more. This was the root cause of the SK041H
# bug: one shared function tried to strip province codes everywhere, which
# accidentally removed "SK" from a plate that genuinely starts with SK.
#
# Plate reads file:  SK041H          (already clean, just uppercase + trim)
# Payments file:     ="SK041H"       (Excel formula wrapper, strip it off)
# Citations file:    SK-SK041H-NA    (PROVINCE-PLATE-SUFFIX, extract middle)
# =============================================================================

def normalise_reads_plate(raw) -> str:
    """
    Plate reads are already clean — just standardise case and whitespace.

    pd.isna() catches empty cells (pandas stores them as NaN = Not a Number).
    We return "" so callers can filter these rows out with  df[df["plate"] != ""]
    """
    if pd.isna(raw):
        return ""
    return str(raw).strip().upper()


def normalise_payments_plate(raw) -> str:
    """
    Payments plates are wrapped in an Excel formula-quoting syntax: ="SK041H"
    We strip the leading =" and the trailing " — nothing else.

    Why lstrip/rstrip instead of a simple replace?
    lstrip('="') removes any combination of = and " from the LEFT end only.
    rstrip('"')  removes " from the RIGHT end only.
    This avoids accidentally removing characters from the middle of a plate.
    """
    if pd.isna(raw):
        return ""
    s = str(raw).strip().upper()
    if s.startswith('="') and s.endswith('"'):
        return s[2:-1]          # slice off first two and last one character
    return s.lstrip('="').rstrip('"')


def normalise_citations_plate(raw) -> str:
    """
    Citations plates follow the format:  PROVINCE-PLATE-SUFFIX
    Examples:  BC-SK041H-NA    AB-XR099L-NA    SK-SK041H-NA

    We split on "-" and take the middle segment (index 1).

    Edge cases handled:
      "LICENSE #"      — the export reprints its column header every page.
                         Return "" to filter it out.
      "BC-  -NA"       — blank plate in the middle. Return "" to skip.
      "BC-XE115F-BIKE" — non-standard suffix. Still works: we always take [1].
      "BC-SK-041H-NA"  — 4 parts instead of 3. Join the middle segments.
      Anything else    — malformed. Return "" to skip safely.
    """
    if pd.isna(raw):
        return ""

    s = str(raw).strip().upper()

    if not s or s == "LICENSE #":
        return ""

    parts = [p.strip() for p in s.split("-")]

    if len(parts) == 3:
        return parts[1]                         # normal case: take middle

    if len(parts) > 3:
        return "".join(parts[1:-1])             # join everything between first and last

    if len(parts) == 1:
        return parts[0]                         # already a bare plate, use as-is

    return ""                                   # anything else: skip


# =============================================================================
# SECTION 2 — FILE UTILITY HELPERS
# =============================================================================
#
# Small reusable functions that don't belong to any one file format.
# Keeping them here means the loaders below stay focused and readable.
# =============================================================================

def excel_engine_for(path: str) -> str | None:
    """
    Returns the right pandas engine for the given file extension.

    .xls  (old Excel format) needs xlrd   — openpyxl can't read it.
    .xlsx (new Excel format) needs openpyxl — xlrd can't read it.

    Path(path).suffix gives us the file extension including the dot.
    .lower() handles "FILE.XLS" the same as "file.xls".
    """
    suffix = Path(path).suffix.lower()
    if suffix == ".xls":
        return "xlrd"
    if suffix == ".xlsx":
        return "openpyxl"
    return None


def find_column(df: pd.DataFrame, keywords: list[str]) -> str | None:
    """
    Finds the first column whose name contains any of the given keywords
    (case-insensitive). Returns None if nothing matches.

    Why this instead of df["License Plate"]?
    Column names sometimes change slightly between exports — "License Plate"
    vs "Licence Plate" vs "LicensePlate". This lets us find the right column
    regardless of minor naming differences.

    Example:
        find_column(df, ["license plate", "licence plate", "plate"])
        # returns "License Plate" if that column exists
    """
    return next(
        (col for col in df.columns
         if any(k in str(col).lower() for k in keywords)),
        None
    )


def require_columns(df: pd.DataFrame, required: list[str], file_label: str) -> None:
    """
    Raises a clear ValueError if any expected columns are missing.

    Without this, a missing column produces a confusing KeyError deep inside
    the code. This surfaces the problem immediately with a useful message.
    """
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(
            f"{file_label} is missing column(s): {missing}\n"
            f"Available columns: {list(df.columns)}"
        )


# =============================================================================
# SECTION 3 — FILE LOADERS
# =============================================================================
#
# One function per file. Each one:
#   1. Reads the raw file into a DataFrame
#   2. Handles that file's specific format quirks (header rows, engines, etc.)
#   3. Normalises the plate column using the file-specific function above
#   4. Returns a minimal, clean DataFrame with only the columns we need
#
# Keeping loaders separate means a format change next month only touches
# one function — nothing else in the script breaks.
# =============================================================================

def _find_reads_header_row(path: str, engine: str | None) -> int:
    """
    Scans the first 10 rows to find which row contains the plate column header.

    Two known formats:
      Old format (Cloudrunner):  header on row 2  (index 1)
        - Row 1 is a report timestamp line
        - Columns: "Plate number", "Local time (PDT)", ...
      New format (Genetec):      header on row 7  (index 6)
        - Rows 1-6 are report metadata (Report name, user, date, count, range)
        - Columns: "Plate read", "Read timestamp", ...

    We scan instead of hardcoding so this keeps working if the metadata
    row count changes in future exports.
    """
    df_raw = pd.read_excel(path, header=None, engine=engine)
    plate_keywords = {"plate read", "plate number"}
    for i in range(min(10, len(df_raw))):
        row_vals = {str(v).strip().lower() for v in df_raw.iloc[i] if pd.notna(v)}
        if row_vals & plate_keywords:
            return i
    return 1   # safe fallback — old format default


def load_plate_reads(path: str) -> pd.DataFrame:
    """
    Loads plate reads from either the old Cloudrunner format or the new
    Genetec format. The two differ in header row position and column names:

      Old (Cloudrunner):
        header row:   index 1  (row 2)
        plate column: "Plate number"
        time column:  "Local time (PDT)"  — stored as a string
                      e.g. "04/30/2026, 11:56:09 PM"

      New (Genetec):
        header row:   index 6  (row 7)
        plate column: "Plate read"
        time column:  "Read timestamp"    — stored as a native Excel datetime
                      e.g. 2026-04-01 10:54:22.023 (pandas reads as Timestamp)

    Both formats end up as a clean DataFrame with columns [plate, local_time].
    """
    engine = excel_engine_for(path)
    header_row = _find_reads_header_row(path, engine)

    df = pd.read_excel(path, header=header_row, engine=engine)
    df.columns = df.columns.str.strip()

    # Flexible column detection — works with both old and new column names.
    # Keywords are ordered most-specific first so "plate read" matches before
    # the generic "plate" fallback.
    plate_col = find_column(df, ["plate read", "plate number", "plate"])
    time_col  = find_column(df, ["read timestamp", "local time"])

    if not plate_col:
        raise ValueError(
            f"Plate reads file has no recognisable plate column.\n"
            f"Available columns: {list(df.columns)}"
        )
    if not time_col:
        raise ValueError(
            f"Plate reads file has no recognisable timestamp column.\n"
            f"Available columns: {list(df.columns)}"
        )

    df = df.rename(columns={plate_col: "plate_raw", time_col: "local_time"})

    df["plate"] = df["plate_raw"].apply(normalise_reads_plate)
    df = df[df["plate"] != ""]

    # Timestamp parsing — detect format from the first non-null value.
    #
    # Old format: the column contains strings like "04/30/2026, 11:56:09 PM".
    #   pd.to_datetime with errors="coerce" alone won't parse this reliably,
    #   so we use the explicit format string.
    #
    # New format: pandas already read the column as Timestamp objects because
    #   Excel stores them as native datetime values. pd.to_datetime on a
    #   Series of Timestamps is a no-op — it just confirms the type.
    sample = df["local_time"].dropna().iloc[0] if not df["local_time"].dropna().empty else None

    if isinstance(sample, str):
        # Old Cloudrunner string format
        df["local_time"] = pd.to_datetime(
            df["local_time"],
            format="%m/%d/%Y, %I:%M:%S %p",
            errors="coerce",
        )
    else:
        # New Genetec native datetime format (or anything pandas already parsed)
        df["local_time"] = pd.to_datetime(df["local_time"], errors="coerce")

    df = df.dropna(subset=["local_time"])

    # Deduplicate same plate + same timestamp.
    # Two cameras at the same entrance can fire simultaneously.
    df = df.drop_duplicates(subset=["plate", "local_time"], keep="first")

    return df[["plate", "local_time"]].copy()


def load_payments(path: str) -> pd.DataFrame:
    """
    Loads the payments CSV.

    dtype=str reads every column as plain text. Without this, pandas might
    try to parse "="SK041H"" as a formula, or turn numeric-looking plates
    like "0006" into the integer 6.
    """
    df = pd.read_csv(path, dtype=str)
    df.columns = df.columns.str.strip()

    plate_col = find_column(df, ["license plate", "licence plate", "plate"])
    if not plate_col:
        raise ValueError(
            f"Payments file has no recognisable plate column.\n"
            f"Available columns: {list(df.columns)}"
        )

    df["plate"] = df[plate_col].apply(normalise_payments_plate)
    df = df[df["plate"] != ""]

    # We only need to know whether a plate paid at all — not how many times.
    # drop_duplicates() keeps one row per unique plate.
    return df[["plate"]].drop_duplicates().copy()


def load_citations(path: str) -> pd.DataFrame:
    """
    Loads the citations file (.xls or .xlsx).

    FORMAT QUIRKS:
      1. Legacy .xls format requires engine="xlrd".
      2. The first 9 rows are report metadata. Real headers are on row 10
         (index 9 in zero-based counting), so we use header=9.
    """
    engine = excel_engine_for(path)

    try:
        df = pd.read_excel(path, engine=engine, header=9)
    except ImportError:
        raise ImportError(
            "Reading .xls files requires xlrd.\n"
            "Fix: pip install xlrd"
        )

    df.columns = df.columns.str.strip()

    plate_col = find_column(df, ["license #", "licence #", "license", "licence", "plate"])
    if not plate_col:
        raise ValueError(
            f"Citations file has no recognisable plate column.\n"
            f"Available columns: {list(df.columns)}"
        )

    df["plate"] = df[plate_col].apply(normalise_citations_plate)
    df = df[df["plate"] != ""]

    return df[["plate"]].drop_duplicates().copy()


def load_permits(path: str) -> pd.DataFrame:
    """
    Loads the optional permit holders file (when available).

    Flexible column detection handles different export formats — the permit
    system may label columns differently from the payments system.
    """
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(path, dtype=str)
    else:
        df = pd.read_excel(path, dtype=str, engine=excel_engine_for(path))

    df.columns = df.columns.str.strip()

    plate_col  = find_column(df, ["plate", "licence", "license"])
    permit_col = find_column(df, ["permit"])
    name_col   = find_column(df, ["name"])

    if not plate_col:
        raise ValueError(
            f"Permits file has no recognisable plate column.\n"
            f"Available columns: {list(df.columns)}"
        )

    out = pd.DataFrame()
    out["plate"] = df[plate_col].apply(normalise_payments_plate)

    if permit_col:
        out["permit_number"] = df[permit_col].fillna("").astype(str)

    if name_col:
        out["name"] = df[name_col].fillna("").astype(str)

    out = out[out["plate"] != ""]
    return out.drop_duplicates(subset=["plate"]).copy()


# =============================================================================
# SECTION 4 — VISIT LOGIC (THE CORE)
# =============================================================================
#
# WHAT COUNTS AS A QUALIFYING VISIT:
#   A plate must appear on a given calendar day, AND the time between its
#   first camera read and last camera read that day must be >= min_hours.
#
#   Why first-to-last read per day?
#   The camera system doesn't give entry/exit events — just drive-by reads.
#   The span from first to last read is the best observable proxy for
#   "how long was this car on campus today."
#
# WHAT COUNTS AS QUALIFYING FOR THE PERK:
#   A plate needs min_visits days where it had a qualifying visit.
#   The default is 10 days in the month.
#
# WHY NOT SESSIONS?
#   We considered gap-based session detection but the daily method is
#   simpler, unambiguous, and directly matches the stated rule:
#   "10 separate occasions." A calendar day is a clear, defensible unit.
# =============================================================================

def compute_qualifying_visits(
    reads: pd.DataFrame,
    min_visits: int,
    min_hours: float,
) -> pd.DataFrame:
    """
    Returns one row per plate that has enough qualifying days.

    Parameters:
        reads       — DataFrame with columns: plate, local_time
        min_visits  — number of qualifying days required (default: 10)
        min_hours   — minimum first-to-last span per day (default: 1.0)

    Returns columns: plate, qualifying_days, avg_hours_per_visit
    """

    # STEP A: Extract the calendar date from each timestamp.
    # .dt.date strips the time component: 2026-04-15 09:14:32 -> 2026-04-15
    reads = reads.copy()
    reads["visit_date"] = reads["local_time"].dt.date

    # STEP B: For each plate + date combination, find the first and last read.
    # groupby() splits the table into groups (one per unique plate+date pair).
    # agg() then computes summary values for each group:
    #   "min" of local_time = earliest read = when they first appeared
    #   "max" of local_time = latest read   = when they last appeared
    daily = (
        reads
        .groupby(["plate", "visit_date"])
        .agg(
            first_read=("local_time", "min"),
            last_read= ("local_time", "max"),
        )
        .reset_index()
        # reset_index() brings plate and visit_date back as regular columns
        # rather than index levels, which makes further operations easier
    )

    # STEP C: Calculate how long each daily span was.
    # Subtracting two datetime columns gives a timedelta (a duration).
    # .dt.total_seconds() / 3600 converts that to decimal hours.
    daily["duration_hrs"] = (
        daily["last_read"] - daily["first_read"]
    ).dt.total_seconds() / 3600

    # STEP D: Keep only days where the span met the minimum.
    qualifying = daily[daily["duration_hrs"] >= min_hours].copy()

    if qualifying.empty:
        return pd.DataFrame(columns=["plate", "qualifying_days", "avg_hours_per_visit"])

    # STEP E: Count qualifying days and average duration per plate.
    # nunique() counts distinct values — so if a plate appeared on
    # April 3 in two rows (shouldn't happen after groupby, but safe),
    # it still only counts as 1 day.
    summary = (
        qualifying
        .groupby("plate")
        .agg(
            qualifying_days=    ("visit_date",    "nunique"),
            avg_hours_per_visit=("duration_hrs",  "mean"),
        )
        .reset_index()
    )

    summary["avg_hours_per_visit"] = summary["avg_hours_per_visit"].round(1)

    # STEP F: Keep only plates that hit the minimum visit count.
    return summary[summary["qualifying_days"] >= min_visits].copy()


# =============================================================================
# SECTION 5 — REPORT WRITER
# =============================================================================
#
# Style constants live here at module level so changing the colour scheme
# means editing one place, not hunting through the function.
#
# All colours are hex strings (same format as HTML/CSS):
#   "1F4E79" = dark navy   "FFFFFF" = white
#   "EBF3FB" = light blue  "E2EFDA" = light green (permit holder rows)
#   "595959" = mid grey    "375623" = dark green (permit holder text)
# =============================================================================

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(name="Arial", bold=True, size=14,  color="1F4E79")
BODY_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="EBF3FB")   # even rows
GREEN_FILL  = PatternFill("solid", start_color="E2EFDA")   # permit holders
GREEN_FONT  = Font(name="Arial", size=10, color="375623")
THIN_SIDE   = Side(style="thin", color="B8CCE4")
CELL_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE,
    top=THIN_SIDE,  bottom=THIN_SIDE,
)


def _style_header_row(ws, row_num: int, num_cols: int) -> None:
    """Applies the dark blue header style to every cell in a row."""
    for col in range(1, num_cols + 1):
        cell           = ws.cell(row=row_num, column=col)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.border    = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_report(
    qualifiers: pd.DataFrame,
    output_path: str,
    month_label: str,
    summary_stats: dict,
) -> None:
    """
    Writes the Excel report with two sheets:
      Sheet 1 — "Qualifiers":          one formatted row per winning plate
      Sheet 2 — "Processing Summary":  funnel counts showing how many plates
                                        passed each stage (audit trail)
    """

    # Workbook() creates an empty Excel file in memory.
    # Nothing hits disk until wb.save() at the end.
    wb = Workbook()

    # -------------------------------------------------------------------------
    # SHEET 1 — QUALIFIERS
    # -------------------------------------------------------------------------
    ws = wb.active
    ws.title      = "Parking Perks Qualifiers"
    ws.freeze_panes = "A4"      # rows 1-3 stay visible when scrolling down

    # Row 1: title spanning all columns
    ws.merge_cells("A1:G1")
    ws["A1"]           = f"Parking Perks — Qualifying Participants  |  {month_label}"
    ws["A1"].font      = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    # Row 2: run metadata
    ws.merge_cells("A2:G2")
    ws["A2"] = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}    "
        f"Total qualifiers: {len(qualifiers)}    "
        f"Criteria: {summary_stats['min_visits']}+ days of "
        f"{summary_stats['min_hours']}+ hr  |  Valid payment/permit  |  No citations"
    )
    ws["A2"].font      = Font(name="Arial", size=9, italic=True, color="595959")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # Row 3: column headers
    headers = [
        "#",
        "Plate Number",
        "Name",
        "Permit Number",
        "Days on Campus",
        "Avg Stay (hrs)",
        "Payment Source",
    ]
    for col_num, text in enumerate(headers, 1):
        ws.cell(row=3, column=col_num, value=text)
    _style_header_row(ws, row_num=3, num_cols=len(headers))
    ws.row_dimensions[3].height = 28

    # Rows 4+: one row per qualifying plate
    # Columns 1, 5, 6 contain numbers — centre those; left-align text columns.
    CENTRE_COLS = {1, 5, 6}

    for display_num, (_, row) in enumerate(qualifiers.iterrows(), start=1):
        excel_row = display_num + 3     # offset past the 3 header rows

        row_data = [
            display_num,
            row["plate"],
            row.get("name", ""),
            row.get("permit_number", ""),
            row["qualifying_days"],
            row["avg_hours_per_visit"],
            row.get("payment_source", "Payment"),
        ]

        is_permit = row.get("payment_source") == "Permit"
        row_fill  = GREEN_FILL if is_permit else (ALT_FILL if display_num % 2 == 0 else PatternFill())
        row_font  = GREEN_FONT if is_permit else BODY_FONT

        for col_num, value in enumerate(row_data, 1):
            cell           = ws.cell(row=excel_row, column=col_num, value=value)
            cell.font      = row_font
            cell.fill      = row_fill
            cell.border    = CELL_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_num in CENTRE_COLS else "left",
                vertical="center",
            )
        ws.row_dimensions[excel_row].height = 20

    # Column widths (in character units)
    for col_num, width in enumerate([5, 16, 28, 18, 18, 16, 16], 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # -------------------------------------------------------------------------
    # SHEET 2 — PROCESSING SUMMARY
    # -------------------------------------------------------------------------
    ws2 = wb.create_sheet("Processing Summary")
    ws2.column_dimensions["A"].width = 44
    ws2.column_dimensions["B"].width = 28

    ws2["A1"]      = "Parking Perks — Processing Summary"
    ws2["A1"].font = TITLE_FONT
    ws2.row_dimensions[1].height = 30

    # Each tuple is (label, value). (None, None) inserts a blank spacer row.
    summary_rows = [
        ("Month",    month_label),
        ("Run date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        (None, None),
        ("-- Input files --", None),
        ("Reads file date range",               summary_stats["date_range"]),
        ("Unique dates in reads file",          summary_stats["coverage_days"]),
        ("Cleaned plate read rows",             summary_stats["read_rows"]),
        ("Unique plates in reads",              summary_stats["total_plates"]),
        ("Unique paid plates",                  summary_stats["payment_plates"]),
        ("Unique cited plates",                 summary_stats["citation_plates"]),
        ("Unique permit plates",                summary_stats["permit_plates"]),
        (None, None),
        ("-- Stage 1: Visit behaviour --", None),
        (f"Minimum qualifying days required",   summary_stats["min_visits"]),
        ("Minimum hours per day (first→last)",  summary_stats["min_hours"]),
        ("Plates passing stage 1",              summary_stats["stage1"]),
        (None, None),
        ("-- Stage 2: Valid payment / permit --", None),
        ("Matched via payment",                 summary_stats["stage2_payment"]),
        ("Matched via permit",                  summary_stats["stage2_permit"]),
        ("Plates passing stage 2",              summary_stats["stage2_total"]),
        (None, None),
        ("-- Stage 3: No citations --", None),
        ("Plates removed (had citation)",       summary_stats["removed_citations"]),
        ("FINAL QUALIFIERS",                    summary_stats["final"]),
    ]

    for row_idx, (label, value) in enumerate(summary_rows, start=2):
        is_section = bool(label and label.startswith("--"))
        is_final   = label == "FINAL QUALIFIERS"

        label_cell      = ws2.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(
            name="Arial", size=10,
            bold=(is_section or is_final),
            color="1F4E79" if is_final else "000000",
        )

        if value is not None:
            val_cell           = ws2.cell(row=row_idx, column=2, value=value)
            val_cell.font      = Font(
                name="Arial", size=10,
                bold=is_final,
                color="1F4E79" if is_final else "000000",
            )
            val_cell.alignment = Alignment(horizontal="right")

    # Warn if the reads file is too short to produce official qualifiers
    if summary_stats["coverage_days"] < summary_stats["min_visits"]:
        warn_row = len(summary_rows) + 4
        ws2.merge_cells(
            start_row=warn_row, start_column=1,
            end_row=warn_row,   end_column=2,
        )
        cell       = ws2.cell(row=warn_row, column=1)
        cell.value = (
            f"WARNING: reads file only covers {summary_stats['coverage_days']} unique dates "
            f"but the threshold is {summary_stats['min_visits']}. "
            f"Run with --min-visits {summary_stats['coverage_days']} to test."
        )
        cell.font      = Font(name="Arial", size=9, bold=True, color="7F6000")
        cell.fill      = PatternFill("solid", start_color="FFF2CC")
        cell.alignment = Alignment(wrap_text=True)
        ws2.row_dimensions[warn_row].height = 42

    ws2.cell(
        row=len(summary_rows) + 6, column=1,
        value="Note: green rows on the Qualifiers sheet = permit holders.",
    ).font = Font(name="Arial", size=9, italic=True, color="595959")

    wb.save(output_path)
    print(f"  Report saved: {output_path}")


# =============================================================================
# SECTION 6 — MAIN
# =============================================================================
#
# The conductor. Doesn't process data itself — calls specialists in order
# and passes results along. If you want to understand the flow, read this
# function top to bottom and follow the function calls outward.
# =============================================================================

def main() -> None:

    # ---- ARGUMENT PARSING --------------------------------------------------
    parser = argparse.ArgumentParser(
        description="Parking Perks — monthly qualifying plate report",
    )
    parser.add_argument("--reads",      required=True,
                        help="Plate reads file (.xlsx)")
    parser.add_argument("--payments",   required=True,
                        help="Payments file (.csv)")
    parser.add_argument("--citations",  required=True,
                        help="Citations file (.xls or .xlsx)")
    parser.add_argument("--permits",    default=None,
                        help="Optional: permit holders file (.csv or .xlsx)")
    parser.add_argument("--output",     default="Parking_Perks_Report.xlsx",
                        help="Output filename (default: Parking_Perks_Report.xlsx)")
    parser.add_argument("--min-visits", type=int,   default=10,
                        help="Qualifying days required (default: 10)")
    parser.add_argument("--min-hours",  type=float, default=1.0,
                        help="Minimum hours on campus per day (default: 1.0)")
    args = parser.parse_args()

    # Build a readable month label from the reads filename.
    # "April_Plate_Reads.xlsx" -> "April Plate Reads"
    month_label = (
        os.path.basename(args.reads)
        .replace("_", " ")
        .replace(".xlsx", "")
        .replace(".xls",  "")
    )

    # ---- LOAD FILES --------------------------------------------------------
    print("\nLoading and cleaning files...")
    reads     = load_plate_reads(args.reads)
    payments  = load_payments(args.payments)
    citations = load_citations(args.citations)
    permits   = (
        load_permits(args.permits)
        if args.permits
        else pd.DataFrame(columns=["plate"])
    )

    coverage_days = reads["local_time"].dt.date.nunique()
    date_min      = reads["local_time"].dt.date.min()
    date_max      = reads["local_time"].dt.date.max()
    date_range    = f"{date_min} to {date_max}"
    total_plates  = reads["plate"].nunique()

    # :, formats numbers with commas: 50000 -> "50,000"
    print(f"  Plate reads:  {len(reads):,} rows | {total_plates:,} unique plates")
    print(f"  Date range:   {date_range}  ({coverage_days} unique dates)")
    print(f"  Payments:     {len(payments):,} unique plates")
    print(f"  Citations:    {len(citations):,} unique plates")
    if not permits.empty:
        print(f"  Permits:      {len(permits):,} unique plates")

    if coverage_days < args.min_visits:
        print(
            f"\n  WARNING: reads file covers {coverage_days} dates "
            f"but threshold is {args.min_visits}. "
            f"Try --min-visits {coverage_days} to test the pipeline."
        )

    # ---- STAGE 1: VISIT BEHAVIOUR ------------------------------------------
    print(f"\nStage 1 — plates with {args.min_visits}+ days of {args.min_hours}+ hr...")
    stage1 = compute_qualifying_visits(reads, args.min_visits, args.min_hours)
    print(f"  -> {len(stage1):,} plates qualify")

    # ---- STAGE 2: VALID PAYMENT OR PERMIT ----------------------------------
    # Python sets give O(1) lookup — "is this plate in the set" is instant
    # regardless of how large the set is. Lists get slower as they grow.
    paid_plates   = set(payments["plate"])
    permit_plates = set(permits["plate"]) if not permits.empty else set()

    # .isin(set) produces a True/False column: True where the plate is in the set
    stage1["has_payment"]   = stage1["plate"].isin(paid_plates)
    stage1["has_permit"]    = stage1["plate"].isin(permit_plates)
    stage1["valid_payment"] = stage1["has_payment"] | stage1["has_permit"]   # OR

    stage2_payment = int(stage1["has_payment"].sum())
    stage2_permit  = int(stage1["has_permit"].sum())
    stage2         = stage1[stage1["valid_payment"]].copy()

    print(f"\nStage 2 — cross-referencing payments and permits...")
    print(f"  Matched via payment: {stage2_payment}")
    print(f"  Matched via permit:  {stage2_permit}")
    print(f"  -> {len(stage2):,} plates have valid payment or permit")

    # ---- STAGE 3: NO CITATIONS ---------------------------------------------
    cited_plates = set(citations["plate"])

    # ~ inverts a boolean column: True -> False, False -> True
    # So ~stage2["plate"].isin(cited_plates) means "plate is NOT cited"
    stage3        = stage2[~stage2["plate"].isin(cited_plates)].copy()
    removed_count = len(stage2) - len(stage3)

    print(f"\nStage 3 — removing plates with citations...")
    print(f"  Removed: {removed_count}")
    print(f"  -> {len(stage3):,} final qualifiers")

    # ---- ENRICH WITH PERMIT DATA -------------------------------------------
    # Label each qualifier's payment source for the report
    stage3["payment_source"] = stage3.apply(
        lambda r: "Permit" if r["has_permit"] else "Payment",
        axis=1,     # axis=1 = apply function row-by-row
    )

    # Merge in names and permit numbers if we have them.
    # how="left" keeps all rows from stage3; adds permit columns where matched,
    # fills NaN where there is no permit record for that plate.
    if not permits.empty:
        permit_cols = ["plate"]
        if "name"          in permits.columns: permit_cols.append("name")
        if "permit_number" in permits.columns: permit_cols.append("permit_number")
        stage3 = stage3.merge(permits[permit_cols], on="plate", how="left")

    if "name"          not in stage3.columns: stage3["name"]          = ""
    if "permit_number" not in stage3.columns: stage3["permit_number"] = ""
    stage3[["name", "permit_number"]] = stage3[["name", "permit_number"]].fillna("")

    # Sort by most days first — top performers appear at the top of the report
    stage3 = stage3.sort_values("qualifying_days", ascending=False).reset_index(drop=True)

    # ---- WRITE REPORT ------------------------------------------------------
    summary_stats = {
        "date_range":        date_range,
        "coverage_days":     coverage_days,
        "read_rows":         len(reads),
        "total_plates":      total_plates,
        "payment_plates":    len(payments),
        "citation_plates":   len(citations),
        "permit_plates":     len(permits) if not permits.empty else 0,
        "min_visits":        args.min_visits,
        "min_hours":         args.min_hours,
        "stage1":            len(stage1),
        "stage2_payment":    stage2_payment,
        "stage2_permit":     stage2_permit,
        "stage2_total":      len(stage2),
        "removed_citations": removed_count,
        "final":             len(stage3),
    }

    print(f"\nWriting report...")
    write_report(stage3, args.output, month_label, summary_stats)
    print(f"  Done. Final qualifier count: {len(stage3)}\n")


# =============================================================================
# ENTRY POINT
# =============================================================================
#
# This block only runs when you execute the file directly:
#   python parking_perks.py --reads ...
#
# If another Python file imports this one (e.g. for testing), this block
# is skipped — the import won't accidentally trigger the whole pipeline.
# =============================================================================

if __name__ == "__main__":
    main()